# Created by: PyQt5 UI code generator 5.15.7
import threading
import os
import sys
import pandas as pd
import openpyxl
import glob
import shutil
import numpy as np
import configparser
import matplotlib.pyplot as plt
import plotly.graph_objs as go
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtGui import QMovie
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import NumberFormat
from mpl_toolkits.mplot3d import Axes3D

        
class Ui_Measurements(object):
    def setupUi(self, Measurements):
        Measurements.setObjectName("Measurements")
        Measurements.resize(600, 320)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("logo-orion-vertical-1.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        Measurements.setWindowIcon(icon)
        Measurements.setStyleSheet("")
        self.centralwidget = QtWidgets.QWidget(Measurements)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setContentsMargins(0, 0, 0, 0)
        self.gridLayout.setSpacing(0)
        self.gridLayout.setObjectName("gridLayout")
        self.widget = QtWidgets.QWidget(self.centralwidget)
        self.widget.setStyleSheet("background-color:rgb(0, 66, 107)")
        self.widget.setObjectName("widget")
        self.folder_select = QtWidgets.QPushButton(self.widget)
        self.folder_select.setGeometry(QtCore.QRect(25, 40, 550, 20))
        self.folder_select.setStyleSheet("background-color: rgb(0, 105, 165);\n"
                                        "color: rgb(255,255, 255);\n"
                                        "font: 75 10pt \"MS Shell Dlg 2\";")
        self.folder_select.setText("Select folder")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(275, 140, 100, 100))
        self.label.setMinimumSize(QtCore.QSize(100, 100))
        self.label.setMaximumSize(QtCore.QSize(100, 100))
        self.label.setObjectName("label")
        self.buttonBox = QtWidgets.QDialogButtonBox(self.widget)
        self.buttonBox.setGeometry(QtCore.QRect(175, 120, 200, 30))
        self.buttonBox.setStyleSheet("background-color: rgb(0, 105, 165);\n"
                                    "color: rgb(255, 255, 255);\n"
                                    "font: 75 10pt \"MS Shell Dlg 2\";")
        self.buttonBox.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel|QtWidgets.QDialogButtonBox.Ok)
        self.buttonBox.setObjectName("buttonBox")
        self.lineEdit = QtWidgets.QLineEdit(self.widget)
        self.lineEdit.setGeometry(QtCore.QRect(200, 250, 200, 20))
        self.lineEdit.setStyleSheet("color: rgb(255, 255, 255);")
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit.setReadOnly(True)
        self.gridLayout.addWidget(self.widget, 0, 0, 1, 1)
        self.checkBox = QtWidgets.QCheckBox(self.centralwidget)
        self.checkBox.setGeometry(QtCore.QRect(20, 80, 82, 17))
        self.checkBox.setStyleSheet("color: rgb(255, 255, 255);\n"
                                    "font: 75 10pt \"MS Shell Dlg 2\"; font-weight: bold")
        self.checkBox.setObjectName("checkBox")
        self.checkBox_1 = QtWidgets.QCheckBox(self.centralwidget)
        self.checkBox_1.setGeometry(QtCore.QRect(120, 80, 82, 17))
        self.checkBox_1.setStyleSheet("color: rgb(255, 255, 255);\n"
                                        "font: 75 10pt \"MS Shell Dlg 2\"; font-weight: bold")
        self.checkBox_1.setObjectName("checkBox_1")
        self.checkBox_2 = QtWidgets.QCheckBox(self.centralwidget)
        self.checkBox_2.setGeometry(QtCore.QRect(220, 80, 111, 17))
        self.checkBox_2.setStyleSheet("color: rgb(255, 255, 255);\n"
                                        "font: 75 10pt \"MS Shell Dlg 2\";font-weight: bold")
        self.checkBox_2.setObjectName("checkBox_2")
        self.checkBox_3 = QtWidgets.QCheckBox(self.centralwidget)
        self.checkBox_3.setGeometry(QtCore.QRect(320, 80, 82, 17))
        self.checkBox_3.setStyleSheet("color: rgb(255, 255, 255);\n"
                                        "font: 75 10pt \"MS Shell Dlg 2\";font-weight: bold")
        self.checkBox_3.setObjectName("checkBox_3")
        self.checkBox_4 = QtWidgets.QCheckBox(self.centralwidget)
        self.checkBox_4.setGeometry(QtCore.QRect(420, 80, 82, 17))
        self.checkBox_4.setStyleSheet("color: rgb(255, 255, 255);\n"
                                        "font: 75 10pt \"MS Shell Dlg 2\";font-weight: bold")
        self.checkBox_4.setObjectName("checkBox_4")
        self.checkBox_5 = QtWidgets.QCheckBox(self.centralwidget)
        self.checkBox_5.setGeometry(QtCore.QRect(520, 80, 82, 17))
        self.checkBox_5.setStyleSheet("color: rgb(255, 255, 255);\n"
                                        "font: 75 10pt \"MS Shell Dlg 2\";font-weight: bold")
        self.checkBox_5.setObjectName("checkBox_5")
        Measurements.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(Measurements)
        self.menubar.setObjectName("menubar")
        self.menuOption = QtWidgets.QMenu(self.menubar)
        self.menuOption.setObjectName("menuOption")
        self.menuHelp = QtWidgets.QMenu(self.menubar)
        self.menuHelp.setObjectName("menuHelp")
        Measurements.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(Measurements)
        self.statusbar.setObjectName("statusbar")
        Measurements.setStatusBar(self.statusbar)
        self.actionData = QtWidgets.QAction(Measurements)
        self.actionData.setObjectName("actionData")
        self.actionColor = QtWidgets.QAction(Measurements)
        self.actionColor.setObjectName("actionColor")
        self.actionAbout_Minerva = QtWidgets.QAction(Measurements)
        self.actionAbout_Minerva.setObjectName("actionAbout_Minerva")
        self.actionAbout_Measurement = QtWidgets.QAction(Measurements)
        self.actionAbout_Measurement.setObjectName("actionAbout_Measurement")
        self.actionDoc = QtWidgets.QAction(Measurements)
        self.actionDoc.setObjectName("actionDoc")
        self.menuOption.addAction(self.actionData)
        self.menuOption.addAction(self.actionColor)
        self.menubar.addAction(self.menuOption.menuAction())
        self.menuHelp.addAction(self.actionAbout_Measurement)
        self.menuHelp.addSeparator()
        self.menuHelp.addAction(self.actionDoc)
        self.menuHelp.addAction(self.actionAbout_Minerva)
        self.menubar.addAction(self.menuHelp.menuAction())
        self.actionData.triggered.connect(self.open_data)
        self.actionColor.triggered.connect(self.open_color)
        self.actionAbout_Minerva.triggered.connect(self.open_about_minerva)
        self.actionAbout_Measurement.triggered.connect(self.open_about_measurement)
        self.actionDoc.triggered.connect(self.open_doc)

        self.retranslateUi(Measurements)
        QtCore.QMetaObject.connectSlotsByName(Measurements)
    
    def retranslateUi(self, Measurements):
        _translate = QtCore.QCoreApplication.translate
        Measurements.setWindowTitle(_translate("Measurements", "Measurements"))
        self.folder_select.setText(_translate("Measurements", "Select File Folder"))
        self.folder_select.clicked.connect(self.select_folder)
        self.lineEdit.setText(_translate("Measurements", "  Version 1.5 by Minerva Dev® & Orion"))
        self.checkBox.setText(_translate("MainWindow", "HeatMap"))
        self.checkBox_1.setText(_translate("Measurements", "Surface"))
        self.checkBox_2.setText(_translate("Measurements", "ID (mm)"))
        self.checkBox_3.setText(_translate("Measurements", "OoR (mm)"))
        self.checkBox_4.setText(_translate("Measurements", "OoR(%)"))
        self.checkBox_5.setText(_translate("Measurements", "Ova"))
        self.menuOption.setTitle(_translate("Measurements", "Option"))
        self.actionData.setText(_translate("Measurements", "Data"))
        self.actionColor.setText(_translate("Measurements", "Color"))
        self.menuHelp.setTitle(_translate("Measurements", "Help"))
        self.actionAbout_Minerva.setText(_translate("Measurements", "About Minerva"))
        self.actionDoc.setText(_translate("Measurements", "Doc"))
        self.actionAbout_Measurement.setText(_translate("Measurements", "About Measurement"))
        self.buttonBox.rejected.connect(self.close_program)
        self.buttonBox.accepted.connect(self.run_program)


    def open_data(self):
        config = configparser.ConfigParser()
        config.read("config.ini")

        try:
            idn = config.getfloat("DEFAULT", "idn")
            wt = config.getfloat("DEFAULT", "wt")
            cra = config.getfloat("DEFAULT", "cra")
            od = config.getfloat("DEFAULT", "od")
        except (configparser.NoOptionError, ValueError):
            idn = 162.15
            wt = 22.00
            cra = 2.50
            od = 237.30

        data = QtWidgets.QDialog()
        data.setWindowTitle("Data")
        data.resize(320, 400)
        layout = QtWidgets.QVBoxLayout()
        
        # Add image
        ref = QtGui.QPixmap("pipe-and-tube.png")
        label_ref = QtWidgets.QLabel()
        label_ref.setPixmap(ref)
        layout.addWidget(label_ref)
        
        # Add label and line edit for idn
        label = QtWidgets.QLabel("ID nominal")
        layout.addWidget(label)
        line_edit = QtWidgets.QLineEdit()
        line_edit.setText("%.2f" % idn)
        layout.addWidget(line_edit)
        
        # Add label and line edit for od
        label_od = QtWidgets.QLabel("OD")
        layout.addWidget(label_od)
        line_edit_1 = QtWidgets.QLineEdit()
        line_edit_1.setText("%.2f" % od)
        layout.addWidget(line_edit_1)
        
        # Add label and line edit for wt
        label_wt = QtWidgets.QLabel("Wall Thickness")
        layout.addWidget(label_wt)
        line_edit_2 = QtWidgets.QLineEdit()
        line_edit_2.setText("%.2f" % wt)
        layout.addWidget(line_edit_2)
        
        # Add label and line edit for cra
        label_cra = QtWidgets.QLabel("CRA")
        layout.addWidget(label_cra)
        line_edit_3 = QtWidgets.QLineEdit()
        line_edit_3.setText("%.2f" % cra)
        layout.addWidget(line_edit_3)
    
        data.finished.connect(lambda: self.save_idn(line_edit.text(), config))
        data.finished.connect(lambda: self.save_od(line_edit_1.text(), config))
        data.finished.connect(lambda: self.save_wt(line_edit_2.text(), config))
        data.finished.connect(lambda: self.save_cra(line_edit_3.text(), config))
        data.setLayout(layout)
        data.exec_()

    def open_color(self):
        config = configparser.ConfigParser()
        config.read("config.ini")

        try:
            bg = config.get("DEFAULT", "bg")
            word = config.get("DEFAULT", "word")
        except (configparser.NoOptionError, ValueError):
            bg = "224C5A"
            word = "f08a04"

        cores = QtWidgets.QDialog()
        cores.setWindowTitle("Colors")
        cores.resize(320, 400)
        layout = QtWidgets.QVBoxLayout()

        # Add image
        ref = QtGui.QPixmap("bg and word.png")
        label_ref = QtWidgets.QLabel()
        label_ref.setPixmap(ref)
        layout.addWidget(label_ref)

        # Add label with explanation text
        explanation = QtWidgets.QLabel("By default, the background color will be set to 224C5A and the color font to f08a04 for the cells\n"
                                    "in the first row, as shown in the above figure. However, you can modify these colors by using the options\n"
                                    "Background and Color Font below. Once you make changes, the program will create a file with your last\n"
                                    "modification. To change the colors, you need to input the hexadecimal color format without the # symbol.")
        layout.addWidget(explanation)


        # Add label and line edit for bg (Background for excel)
        label_bg = QtWidgets.QLabel("Background Color: backgorund color of cell excel")
        layout.addWidget(label_bg)
        line_edit = QtWidgets.QLineEdit()
        line_edit.setText(bg)
        layout.addWidget(line_edit)
            
        # Add label and line edit for word (color font for words in excel)
        label_word = QtWidgets.QLabel("Color Font: font color of first line of cell excel")
        layout.addWidget(label_word)
        line_edit_1 = QtWidgets.QLineEdit()
        line_edit_1.setText(word)
        layout.addWidget(line_edit_1)
        cores.finished.connect(lambda: self.save_bg(line_edit.text(), config))
        cores.finished.connect(lambda: self.save_word(line_edit_1.text(), config))
        cores.setLayout(layout)
        cores.exec_()


    def save_idn(self, idn, config):
        config.set("DEFAULT", "idn", idn.replace(",", "."))
        with open("config.ini", "w") as config_file:
            config.write(config_file)

    def save_od(self, od, config):
        config.set("DEFAULT", "od", od.replace(",", "."))
        with open("config.ini", "w") as config_file:
            config.write(config_file)

    def save_wt(self, wt, config):
        config.set("DEFAULT", "wt", wt.replace(",", "."))
        with open("config.ini", "w") as config_file:
            config.write(config_file)

    def save_cra(self, cra, config):
        config.set("DEFAULT", "cra", cra.replace(",", "."))
        with open("config.ini", "w") as config_file:
            config.write(config_file)
            
    def save_bg(self, color_string, config):
        config.set("DEFAULT", "bg", color_string)
        with open("config.ini", "w") as configfile:
            config.write(configfile)

    def save_word(self, color_string, config):
        config.set("DEFAULT", "word", color_string)
        with open("config.ini", "w") as configfile:
            config.write(configfile)
        
        
    def open_about_minerva(self):
        about_Minerva = QtWidgets.QDialog()
        about_Minerva.setWindowTitle("About Minerva")
        about_Minerva.resize(280, 420)
        layout = QtWidgets.QVBoxLayout()
        label = QtWidgets.QLabel("Minerva Dev is a company founded in 2023 that specializes in solving software problems\nfor businesses using the Python programming language."
                                 "With an on-demand approach, the company provides customized and efficient solutions to meet each client's specific needs."
                                 "With a highly skilled and experienced team, Minerva Dev is committed to providing quality service, always aiming for customer satisfaction and exceeding expectations.\n\n"
                                 "The name Minerva Dev is inspired by the patron of the institution where the company's founders first met. "
                                 "Founded by friends in 2023, the company is built on a strong foundation of mutual respect and shared values, which are reflected in its commitment"
                                 "to providing excellent service and building lasting relationships with its clients.")
        label.setWordWrap(True)
        layout.addWidget(label)
        about_Minerva.setLayout(layout)
        about_Minerva.exec_()

    def open_doc(self):
        doc = QtWidgets.QDialog()
        doc.setWindowTitle("Doc")
        doc.resize(280, 420)
        layout = QtWidgets.QVBoxLayout()
        label = QtWidgets.QLabel()
        title1 = QtWidgets.QLabel("Errno 13")
        title1.setAlignment(QtCore.Qt.AlignCenter)
        title1.setFont(QtGui.QFont("MS Shell Dlg 2", 12, QtGui.QFont.Bold))
        layout.addWidget(title1)

        para1 = QtWidgets.QLabel("Indicates that you are trying to run the program with an open Excel file in the same folder. This may occur if you have previously executed the program, created an Excel file, and the file is still open when you are trying to run the program again.")
        para1.setWordWrap(True)
        layout.addWidget(para1)

        title2 = QtWidgets.QLabel("WinError 267")
        title2.setAlignment(QtCore.Qt.AlignCenter)
        title2.setFont(QtGui.QFont("MS Shell Dlg 2", 12, QtGui.QFont.Bold))
        layout.addWidget(title2)

        para2 = QtWidgets.QLabel("Indicates that you are trying to run the program with a folder that has already had the program executed and contains at least one figures folder, which could be a single Heat Map folder, ID folder, OoR (mm) folder, OoR (%) folder, or a combination of them. The error message recommends deleting the folder and restarting the program for a new round or cancelling, as the processing of this folder has already been completed.")
        para2.setWordWrap(True)
        layout.addWidget(para2)
        label.setWordWrap(True)
        doc.setLayout(layout)
        doc.exec_()

    def open_about_measurement(self):
        about_measurement = QtWidgets.QDialog()
        about_measurement.setWindowTitle("About Measurement")
        about_measurement.resize(450, 420)
        layout = QtWidgets.QVBoxLayout()
        logo = QtWidgets.QLabel()
        logo.setPixmap(QtGui.QPixmap("logo-orion-vertical-1.png"))
        layout.addWidget(logo)
        separator = QtWidgets.QFrame()
        separator.setFrameShape(QtWidgets.QFrame.HLine)
        layout.addWidget(separator)
        label = QtWidgets.QLabel("Measurement is a program designed to Orion make it easy to take and store measurements."
                                 "With its user-friendly interface, you can perform accurate measurements with just a few clicks.\n\n"
                                 "Version: 1.5\n"
                                 "Author: Marcus Filgueiras and Gustavo Pessanha\n"
                                 "Copyright (c) 2023 Minerva Dev")
        label.setWordWrap(True)
        label.setFrameStyle(QtWidgets.QFrame.HLine | QtWidgets.QFrame.Sunken)
        layout.addWidget(label)
        about_measurement.setLayout(layout)
        about_measurement.exec_()

        
    def __init__(self):
        self.selected_folder = ""

    def select_folder(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        folder = QFileDialog.getExistingDirectory(Measurements, "Select folder", options=options)
        #Reset the text to "Select folder" if the user cancelled the file selection
        if folder == '':
            self.folder_select.setText("Select File Folder")
        else:
            self.folder_select.setText(folder)
            self.selected_folder = folder
    

    def run_program(self):
        if self.selected_folder == "":
            # Show an error message if no folder has been selected
            error_msg = "No folder selected. Please select a folder before running the program."
            msg = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "Error", error_msg)
            msg.setDetailedText("Maybe you didn't select the folder with the files and this error appeared on your screen, in this case it will not start the program, please go back and select the folder to run the program.")
            msg.exec_()
        else:
            # set qmovie as label
            self.movie = QMovie("Reload-1s-50px.gif")
            self.label.setMovie(self.movie)
            self.movie.start()
            # Run the program using the selected folder in a new thread
            program_instance = Program()
            program_instance = Program(selected_folder=self.selected_folder, movie=self.movie, label=self.label, checkBox = self.checkBox, checkBox_1 = self.checkBox_1,checkBox_2 = self.checkBox_2, checkBox_3 = self.checkBox_3, checkBox_4 = self.checkBox_4, checkBox_5 = self.checkBox_5 )
            t = threading.Thread(target=program_instance.programa)
            t.start()

    def close_program(self):
        #Close the main window and exit the application
        Measurements.close()
        sys.exit()

class Program:
    def __init__(self, selected_folder=None, movie=None, label=None, checkBox=None, checkBox_1=None,checkBox_2=None, checkBox_3=None, checkBox_4=None, checkBox_5=None):
        self.selected_folder = selected_folder
        self.count = 0
        self.movie = movie
        self.label = label
        self.checkBox = checkBox
        self.checkBox_1 = checkBox_1
        self.checkBox_2 = checkBox_2
        self.checkBox_3 = checkBox_3
        self.checkBox_4 = checkBox_4
        self.checkBox_5 = checkBox_5
        config = configparser.ConfigParser()
        config.read("config.ini")
        self.idn = config.getfloat("DEFAULT", "idn")
        self.od = config.getfloat("DEFAULT", "od")
        self.wt = config.getfloat("DEFAULT", "wt")
        self.cra = config.getfloat("DEFAULT", "cra")
        self.bg = config.get("DEFAULT", "bg")
        self.word = config.get("DEFAULT", "word")

    def show_error_message(self, error_message):
        message = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Critical, "Error", error_message)
        message.exec_()


    def programa(self):
        idn = self.idn
        od = self.od
        cra = self.cra
        wt = self.wt
        bg = self.bg
        word = self.word
        if self.selected_folder:
            #Empty List
            filenames = []
            folder_names = []
            reference_filenames = []
            # Use glob to search for folders with names containing "pr5_files"
            folders = glob.glob(self.selected_folder +'/**/*set*', recursive=True)
            #only name file
            folders1 = [os.path.basename(folder) for folder in folders]
            
            #Extract the data from the pr5 files and transform it into txt with the data you need
            for folder in folders:
                try:
                    folder_names.append(folder)
                    for file in os.listdir(folder):
                        if file.endswith('.pr5'):
                            reference_filenames.append(file)
                            # Open PR5 file for reading
                            with open(os.path.join(folder, file), 'r') as f_in:
                              # Create a text file name based on the PR5 file name
                                txt_file = file[:-3] + 'txt'
                                filenames.append(txt_file)
                                # Open the text file for writing
                                with open(os.path.join(folder, txt_file), 'w') as f_out:
                                  # Loop over the lines of the PR5 file
                                  for line in f_in:
                                    # Check if the current line starts with the string you want
                                    if line.startswith('[Average Data]'):
                                     # Write all lines from here
                                      break
                                  else:
                                    # If we didn't find the string "[Average Data]", continue to the next iteration
                                    continue
                                  # Write all remaining lines to the text file
                                  for line in f_in:
                                    f_out.write(line.replace('.', ',')[:19]+ '\n')
                except NotADirectoryError:
                    error_message = f"The directory '{folder}' did't can processed with .pr5. If you delete the figures folder, it will run normally."
                    message = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Critical, "Error", error_message)
                    message.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    message.setDefaultButton(QtWidgets.QMessageBox.Ok)
                    message.exec_()
                    
                        
            ##########Pega todos os novos arquivos .txt e move ele para um pasta separada###########################
            # Loop over each folder containing the string "pr5_files"
            current_folder = os.path.basename(os.getcwd())
            for folder in folders:
                # Construa o caminho da pasta de destino
                dest_folder = os.path.join(folder, 'txt_files')
                # Crie a pasta de destino, se ela ainda não existir
                if not os.path.exists(dest_folder):
                    os.makedirs(dest_folder)
                # Loop over each file in the current folder
                for file in os.listdir(folder):
                    # Check if the file is a txt file
                    if file.endswith('.txt'):
                        # Build the source and destination paths for the file
                        src_path = os.path.join(folder, file)
                        dest_path = os.path.join(dest_folder, file)
                        # Move the file to the destination folder
                        shutil.move(src_path, dest_path)     
                    
            ####################### Create a new Excel workbook ########################################
            workbook = openpyxl.Workbook()
            arq  = os.path.basename(self.selected_folder)
            output_file = os.path.join(self.selected_folder, arq + ".xlsx")
            workbook.save(output_file)
            # Load the workbook
            workbook = openpyxl.load_workbook(output_file)
            # Get the first sheet
            sheet = workbook.worksheets[0]
            # Rename the sheet to the name of the first folder
            sheet.title = folders1[0]
            # Iterate over the remaining folders
            for folder in folders1[1:]:
                self.count += 1
                # Create a new sheet with the name of the current folder 
                if folder in workbook.sheetnames:
                    print(f'Sheet {folder} already exists')
                else:
                    # Create a new sheet
                    sheet = workbook.create_sheet(title=folder)      
            # Iterate over all sheets
            for worksheet in workbook.worksheets:
                # Set the width of the first eight columns to 130 px
                for i in 'ABCDEFGHIJK':
                    worksheet.column_dimensions[i].width = 18.59098065
                    # Get the cells in the first row
                    first_row_cells = worksheet[1]
            # Save the workbook
            workbook.save(output_file)
            ###################################################################
            # Set the path to the Excel file
            excel_file = output_file
            txt_folders = []
            for folder in folders:
                # Get the full path of the folder
                full_path = os.path.abspath(folder)
                # Get all the subfolders inside the current folder
                subfolders = os.listdir(full_path)
                # Iterate over the subfolders and check if their name ends with "txt_files"
                for subfolder in subfolders:
                    if subfolder.endswith("txt_files"):
                        # If it does, append the full path of the subfolder to txt_folders
                        txt_folders.append(os.path.join(full_path, subfolder))
                        # If it does, append the full path of the subfolder to txt_folders
                        full_path = os.path.join(full_path, subfolder)
                        if full_path not in txt_folders:
                            txt_folders.append(full_path)
            #################################################################################
                
            # Open the Excel workbook
            workbook = openpyxl.load_workbook(output_file)
            # Check if the subfolder ends with "txt_files"
            if subfolder.endswith("txt_files"):
            # If it does, append the full path of the subfolder to txt_folders
                txt_folders.append(os.path.join(full_path, subfolder))
                
            # Load the workbook
            # Set the path to the Excel file
            excel_file = output_file
            txt_folders = []
            for folder in folders:
                # Get the full path of the folder
                txt_folder = os.path.join(folder, 'txt_files')
                txt_folders.append(txt_folder)
            # Iterate over each folder containing txt files
            for txt_folder, sheet in zip(txt_folders, workbook.worksheets):
                # Iterate over each txt file in the folder
                for txt_file in os.listdir(txt_folder):
                    # Check if the file is a txt file
                    if txt_file.endswith('.txt'):
                        # Set the start_column variable to 1
                        # Build the full path of the txt file
                        txt_file_path = os.path.join(txt_folder, txt_file)
                        # Read the txt file into a Pandas DataFrame
                        df = pd.read_csv(txt_file_path, skiprows=6, sep='\s+', decimal=',', header=None)
                        # Crie um novo dataframe com os nomes das colunas
                        # Check if this is the first txt file in the folder
                        if txt_file == os.listdir(txt_folder)[0]:
                            # Select the first 2 columns of the DataFrame
                            df = df[df.columns[:2]]
                        else:
                            # Drop the first column of the DataFrame
                            df = df.drop(df.columns[0], axis=1) 
                        # Transpose the DataFrame
                        df = df.transpose()
                        # Convert the DataFrame to a list of rows
                        rows = dataframe_to_rows(df, index=False)
                        # Get the starting row and column for writing the data
                        # Set the column names
                        column_names = ['Deph (mm)', 'Orion Reference','Avg ID (mm)','Min ID (mm)', 'Max ID (mm)', 'OoR (mm)', 'Ova (%)', 'Min Radial (mm)','Pos of Min Radial (°)', 'Max Radial (mm)','Pos of Max Radial (°)']
                        # Iterate over the column names and set the values in the first row of the worksheet
                        for i, column_name in enumerate(column_names):   
                            sheet.cell(row=1, column=i+1).value = column_name
                        # Convert the DataFrame to a list of lists
                        rows = df.values.tolist()
                        # Iterate over the rows and append them to the worksheet
                        for row in rows:
                            sheet.append(row)
                            
            sheet = workbook.worksheets[0]
            cells = workbook.worksheets[0]
            
            for sheet in workbook.worksheets:
                #Move A2:WB para L2
                last_row_with_data = sheet.max_row
                sheet.move_range(f'A2:WB{last_row_with_data}', rows=-1, cols=11)
                #Formulas for excel
                for row in sheet.iter_rows(min_row=2):
                    #This code calculates the maximum and minimum sum of pairs of cells
                    max_sum = 0
                    min_sum = float('inf')
                    total_sum = 0
                    counta = 0
                    for i in range(12,312):
                    # selecionar a célula atual e a célula seguinte
                        cell1 = sheet.cell(row=row[0].row, column=i)
                        cell2 = sheet.cell(row=row[0].row, column=i+300)
                        # calcular a soma das células atuais
                        sum = cell1.value + cell2.value
                        # atualizar o máximo da soma se a soma atual for maior que o máximo atual
                        max_sum = max(max_sum, sum)
                        min_sum = min(min_sum, sum)
                        total_sum += sum
                        counta += 1
                    average = total_sum / counta
                    oor = max_sum - min_sum
                    #Set the formula
                    #Set the formula for the Avg ID value
                    row[2].value = average
                    #Set the formula for the Min ID value
                    row[3].value = float("{:.5f}".format(float(f"{min_sum}".replace(",", "."))))
                    #Set the formula for the Max ID value
                    row[4].value = float("{:.5f}".format(float(f"{max_sum}".replace(",", "."))))
                    #Set the formula for the OoR value
                    #row[5].value = f"=(E{row[0].row}-D{row[0].row})"
                    row[5].value = oor
                    #Set the formula for the Ova value
                    row[6].value = (oor/average)*100
                    # Set the formula for the maximum value
                    row[9].value = f"=MAX(L{row[0].row}:WM{row[0].row})"
                    # Set the formula for position degrees the maximum value
                    row[10].value=f"ÍNDICE($1:$1;CORRESP(J{row[0].row};L{row[0].row}:WM{row[0].row};0)+11)"
                    # Set the formula for the minimum value
                    row[7].value = f"=MIN(L{row[0].row}:WM{row[0].row})"
                    # Set the formula for position degrees the minimum value
                    row[8].value=f"ÍNDICE($1:$1;CORRESP(H{row[0].row};L{row[0].row}:WM{row[0].row};0)+11)"
            # initialize a counter variable to keep track of the current filename
            filename_counter = 0
            # iterate through the worksheets
            for sheet in workbook.worksheets:
                # get the number of rows in the sheet
                num_rows = sheet.max_row
                # iterate through the rows in the sheet
                for i in range(1,num_rows):
                    # only insert the filename if it exists in the list
                    if filename_counter < len(reference_filenames):
                        row = i + 1  # increment the row index by 1 to account for the header row
                        filename = reference_filenames[filename_counter]  # get the current filename
                        sheet.cell(row=row, column=2).value = filename  # set the value of the cell in the specified row and column
                        filename_counter += 1  # increment the counter
                    else:
                        # reset the counter to 0 if all of the filenames have been inserted
                        filename_counter = 0
            # Modify the font and fill properties of the cells in the first row of each worksheet in the workbook
            for sheet in workbook.worksheets:
                # Get the cells in the first row
                first_row_cells = sheet[1]
                # Iterate over the cells in the first row and set the font and fill
                for cell in first_row_cells:
                    cell.fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=bg)
                    cell.font = Font(color=word, bold=True)
                    # Set the alignment for all cells in the worksheet
                for row in sheet.rows:
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center') 
            # Save the workbook
            workbook.save(output_file)
            # Load the workbook
            workbook = openpyxl.load_workbook(output_file)


            #HeatMap checkBox 
            if self.checkBox.isChecked():
                # Iterate over the sheet names
                for sheet_name in workbook.sheetnames:
                    fig, axs = plt.subplots(1, 1, figsize=(10,10))
                    sheet = workbook[sheet_name]
                    last_row = sheet.max_row
                    data = [row[12:] for row in sheet.iter_rows(min_row=2, max_row=last_row, values_only=True)]
                    data = np.array(data)
                    image = axs.imshow(data, cmap='viridis', aspect = 2)
                    axs.set_title(sheet_name)
                    axs.set_ylabel("Depth")
                    cbar = plt.colorbar(image, ax=axs, shrink=0.2)
                    cbar.ax.set_xlabel("ID radial(mm)")
                    figures_folder = os.path.join(self.selected_folder, "heat map")
                    if not os.path.exists(figures_folder):
                        os.makedirs(figures_folder)
                    filename_fig = os.path.join(figures_folder, f"{sheet_name}.png")
                    if os.path.exists(filename_fig):
                        os.remove(filename_fig)
                    plt.savefig(filename_fig, bbox_inches='tight')

                    
            #Surface checkBox_1 
            if self.checkBox_1.isChecked():
                # Iterate over the sheet names
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    last_row = sheet.max_row
                    data = [row[12:] for row in sheet.iter_rows(min_row=2, max_row=last_row, values_only=True)]
                    data = np.array(data)
                    fig = go.Figure(data=[go.Surface(z=data, colorscale='viridis')])
                    fig.update_layout(scene=dict(xaxis_title='Angle (°)',
                                                  yaxis_title='Depth',
                                                  zaxis_title='ID radial(mm)'))
                    figures_folder = os.path.join(self.selected_folder, "surface")
                    if not os.path.exists(figures_folder):
                        os.makedirs(figures_folder)
                    filename_fig = os.path.join(figures_folder, f"{sheet_name}.html")
                    if os.path.exists(filename_fig):
                        os.remove(filename_fig)
                    fig.write_html(filename_fig, auto_open=False)

            #ID checkBox_2       
            if self.checkBox_2.isChecked():
                #Iterate over the sheet names
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    fig, ax = plt.subplots()
                    last_row = sheet.max_row
                    data1 = [sheet.cell(row=i, column=3).value for i in range(2, last_row + 1)]
                    data2 = [sheet.cell(row=i, column=4).value for i in range(2, last_row + 1)]
                    data3 = [sheet.cell(row=i, column=5).value for i in range(2, last_row + 1)]
                    plt.plot(data1, label="ID avg")
                    plt.plot(data2, label="ID min")
                    plt.plot(data3, label="ID max")
                    ax.yaxis.grid(True)
                    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                    plt.ylabel("ID (mm)")
                    plt.xlabel("Depth " + sheet_name)
                    plt.title("ID", fontweight='bold')
                    ID_folder = os.path.join(self.selected_folder, "ID")
                    if not os.path.exists(ID_folder):
                        os.makedirs(ID_folder)
                    filename_id = os.path.join(ID_folder, f"{sheet_name}.png")
                    if os.path.exists(filename_id):
                        os.remove(filename_id)
                    plt.savefig(filename_id, bbox_inches='tight')

            #OoR (mm) checkBox_3      
            if self.checkBox_3.isChecked():
                #Iterate over the sheet names
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    fig, ax = plt.subplots()
                    last_row = sheet.max_row
                    data4 = [sheet.cell(row=i, column=6).value for i in range(2, last_row + 1)]
                    plt.plot(data4, label="Oor (mm)")
                    ax.yaxis.grid(True)
                    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                    plt.ylabel("OoR(mm)")
                    plt.xlabel("Depth " + sheet_name)
                    plt.title("OoR", fontweight='bold')
                    OoR_folder = os.path.join(self.selected_folder, "OoR (mm)")
                    if not os.path.exists(OoR_folder):
                        os.makedirs(OoR_folder)
                    filename_oor = os.path.join(OoR_folder, f"{sheet_name}.png")
                    if os.path.exists(filename_oor):
                        os.remove(filename_oor)
                    plt.savefig(filename_oor, bbox_inches='tight')

            #OoR(%) checkBox_4      
            if self.checkBox_4.isChecked():
                #Iterate over the sheet names
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    fig, ax = plt.subplots()
                    last_row = sheet.max_row
                    data5 = [sheet.cell(row=i, column=7).value for i in range(2, last_row + 1)]
                    plt.plot(data5, label="Oor (%)")
                    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                    ax.yaxis.grid(True)
                    plt.ylabel("OoR(%)")
                    plt.xlabel("Depth " + sheet_name)
                    plt.title("OoR", fontweight='bold')
                    OoR1_folder = os.path.join(self.selected_folder, "OoR (%)")
                    if not os.path.exists(OoR1_folder):
                        os.makedirs(OoR1_folder)
                    filename_oor1 = os.path.join(OoR1_folder, f"{sheet_name}.png")
                    if os.path.exists(filename_oor1):
                        os.remove(filename_oor1)
                    plt.savefig(filename_oor1, bbox_inches='tight')

            #Ova checkBox_5      
            if self.checkBox_5.isChecked():
                raioid = idn/2
                #Iterate over the sheet names
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for i in range(2, sheet.max_row + 1):
                        linha = []
                        for column in range(12, sheet.max_column + 1):
                            value = sheet.cell(row=i, column=column).value
                            linha.append(float(value))
                        
                        # Create two lists for the x and y coordinates
                        x = np.cos(np.linspace(0, 2*np.pi, len(linha)))
                        y = np.sin(np.linspace(0, 2*np.pi, len(linha)))

                        # Multiply the coordinates by the values in the row
                        x = x * linha
                        y = y * linha

                        # Plot the scatter plot
                        fig, ax = plt.subplots()
                        ax.scatter(x, y, s=0.5)

                        # Add another circle with raioid 
                        circle = plt.Circle((0,0), raioid, color='red', fill=False)
                        ax.add_artist(circle)

                        plt.axis('equal')
                        plt.title("Ova - " + str(sheet.cell(row=i, column=2).value), fontweight='bold')
                        plt.ylabel("mm")
                        plt.xlabel("mm")
                        
                        ova_folder = os.path.join(self.selected_folder, "Ova")
                        if not os.path.exists(ova_folder):
                            os.makedirs(ova_folder)
                        filename_ova = os.path.join(ova_folder, f"{sheet_name}-{sheet.cell(row=i, column=2).value}.png")
                        if os.path.exists(filename_ova):
                            os.remove(filename_ova)
                        plt.savefig(filename_ova, bbox_inches='tight')        

        #Finish program        
        self.movie.stop()
        self.label.setGeometry(292, 140, 76, 10)
        self.label.setStyleSheet("color: white; font-size: 10pt; font-family: MS Shell Dlg 2; font-weight: bold")
        self.label.setText("COMPLETE")
                

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    Measurements = QtWidgets.QMainWindow()
    ui = Ui_Measurements()
    ui.setupUi(Measurements)
    Measurements.show()
    sys.exit(app.exec_())
